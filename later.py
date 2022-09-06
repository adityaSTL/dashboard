import xlrd
import pandas as pd
import numpy as np
from datetime import date,timedelta
import datetime as dt
import openpyxl
import matplotlib.pyplot as plt
from openpyxl_image_loader import SheetImageLoader
import excel2img


####### -1 day  for  script

##setting dates
yesterday_date=pd.to_datetime(date.today()-timedelta(days=3))
today_date=pd.to_datetime(date.today()-timedelta(days=2))
first_date1=pd.to_datetime(date.today()-timedelta(days=31))
print(first_date1)
first_date=pd.to_datetime((date.today()-timedelta(days=2))).replace(day=1)
print("Dates all set")
##Done

##Getting dashboard to write directly
xfile = openpyxl.load_workbook(r'C:\Users\Aditya.gupta\Desktop\automation\Dashboard.xlsx')
sheet = xfile['Sheet1']

 

## Reading old data2
namein=str((date.today()-timedelta(days=3)).strftime('%d-%b-%y'))
namefull= r"C:\Users\Aditya.gupta\Desktop\automation\XL_dated_"+namein+r".xlsx"
wb_obj = openpyxl.load_workbook(namefull)
sheet_obj = wb_obj['Sheet1']
old_GPlitB=sheet_obj.cell(row = 4, column = 5).value
old_PPClearB=sheet_obj.cell(row = 5, column = 5).value
old_GPATB=sheet_obj.cell(row = 6, column = 5).value
old_docB=sheet_obj.cell(row = 7, column = 5).value
old_boqB=sheet_obj.cell(row = 8, column = 5).value
old_GPlitC=sheet_obj.cell(row = 12, column = 5).value
old_PPClearC=sheet_obj.cell(row = 13, column = 5).value
old_GPATC=sheet_obj.cell(row = 14, column = 5).value
old_docC=sheet_obj.cell(row = 15, column = 5).value
old_boqC=sheet_obj.cell(row = 16, column = 5).value
print("Old Data read")
## Dashboard reading excessive complete




##DPR Tracker Package B Reading & correcting errors (date cleaning)
print("Getting pkgB DPR data and cleaning")
namein=str(((date.today()-timedelta(days=2)).strftime('%d-%m-%Y')))
namefull= r"C:\Users\Aditya.gupta\Downloads\Package B _ DPR _"+namein+r" New.xlsb"
loc_pkgB_DPR=(namefull)
df_dprB=pd.read_excel(loc_pkgB_DPR,sheet_name='Daywise Progess Report',skiprows=1)
df_dprB[r' Date of Activity'].replace(r'Final Cumm.',44450,inplace=True)
df_dprB[r' Date of Activity'] = pd.TimedeltaIndex(df_dprB[r' Date of Activity'], unit='d') + dt.datetime(1899, 12, 30)
df_dprB['T&D'] = pd.to_numeric(df_dprB['T&D'], errors='coerce')
df_dprB['T&D'].replace('NaN',0)
df_dprB['DRT'] = pd.to_numeric(df_dprB['DRT'], errors='coerce')
df_dprB['DRT'].replace('NaN',0)
df_dprB['Blowing'] = pd.to_numeric(df_dprB['Blowing'], errors='coerce')
df_dprB['Blowing'].replace('NaN',0)
##Done

##Reading execution plan
print("Reading execution plan Package B & C")
namefull= r"C:\Users\Aditya.gupta\Desktop\automation\Latest Plan.xlsx"
loc_plan=(namefull)
df_plan_B=pd.read_excel(loc_plan,sheet_name='PkgB')
df_plan_C=pd.read_excel(loc_plan,sheet_name='PkgC')
pd.to_datetime(df_plan_B[r'Date'])
pd.to_datetime(df_plan_C[r'Date'])

##Plan manipulation for easier reading for pkg B
df02=df_plan_B['T&D'].groupby(df_plan_B[r'Date']).sum().reset_index()
df04=df_plan_B['DRT'].groupby(df_plan_B[r'Date']).sum().reset_index()
df03=df_plan_B['Blowing'].groupby(df_plan_B[r'Date']).sum().reset_index()
df_0final=pd.concat([df02,df04['DRT'],df03['Blowing']], axis=1)
df_0final=df_0final.set_index(r'Date')
df_0final= df_0final.sort_index()
df_01 = df_0final.loc[first_date1 :today_date].reset_index()

print(df_01)

##Plan manipulation for easier reading for pkg C
df002=df_plan_C['T&D'].groupby(df_plan_C[r'Date']).sum().reset_index()
df004=df_plan_C['DRT'].groupby(df_plan_C[r'Date']).sum().reset_index()
df003=df_plan_C['Blowing'].groupby(df_plan_C[r'Date']).sum().reset_index()
df_00final=pd.concat([df002,df004['DRT'],df003['Blowing']], axis=1)
df_00final=df_00final.set_index(r'Date')
df_00final= df_00final.sort_index()
df_001 = df_00final.loc[first_date1 :today_date].reset_index()

##DPR Tracker Package C Reading
print("Getting pkgC DPR data and cleaning")
namein=str((date.today()-timedelta(days=2)).strftime('%d-%b-%y'))
namefull= r"C:\Users\Aditya.gupta\Downloads\Daily DPR_PKGC FINAL "+namein+r" 1.xlsx"
loc_pkgC_DPR=(namefull)
df_dprC=pd.read_excel(loc_pkgC_DPR,sheet_name='Day Wise Progress',skiprows=1)
pd.to_datetime(df_dprC[r' Date of Activity'])
df_dprC['T&D'] = pd.to_numeric(df_dprC['T&D'], errors='coerce')
df_dprC['T&D'].replace('NaN',0)
df_dprC['DRT'] = pd.to_numeric(df_dprC['DRT'], errors='coerce')
df_dprC['DRT'].replace('NaN',0)
df_dprC['Blowing'] = pd.to_numeric(df_dprC['Blowing'], errors='coerce')
df_dprC['Blowing'].replace('NaN',0)
##done

##POP Tracker Package B DPR Reading
print("Getting pkgB POP data and cleaning")
namein=str((date.today()-timedelta(days=2)).strftime('%d-%b-%Y'))
namefull= r"C:\Users\Aditya.gupta\Downloads\POP DPR_PKG-B@"+namein+r".xlsx"
loc_pkgB_POP=(namefull)
df_POPB=pd.read_excel(loc_pkgB_POP,sheet_name="GP PoP",skiprows=6)
pd.to_datetime(df_POPB[r'GP Lit Up Date'],errors='coerce')

##POP Tracker Package C Reading
print("Getting pkgC POP data and cleaning")
namein=str((date.today()-timedelta(days=2)).strftime('%d-%b-%Y'))
namefull= r"C:\Users\Aditya.gupta\Downloads\POP DPR - PKGC as on "+namein+r".xlsx"
loc_pkgC_POP=(namefull)
df_POPC=pd.read_excel(loc_pkgC_POP,sheet_name="GP PoP",skiprows=5)
df_POPC[r'GP Lit Up Date'] = pd.to_datetime(df_POPC[r'GP Lit Up Date'],errors='coerce')

##AT Tracker Package B Reading
print("Getting pkgB AT data and cleaning")
namein=str((date.today()-timedelta(days=2)).strftime('%d-%b-%y'))
namefull= r"C:\Users\Aditya.gupta\Downloads\PKG-B AT TRACKER "+namein+r".xlsx"
loc_pkgB_AT=(namefull)
df_atB=pd.read_excel(loc_pkgB_AT,sheet_name="Backup Sheet",skiprows=1)

print("Got Pkg B AT Info now getting Doc's submitted and BoQ approval status")
df_docB=pd.read_excel(loc_pkgB_AT,sheet_name="Invoice Tracker Backup",skiprows=1)



##AT Tracker Package C Reading
print("Getting pkgC AT data and cleaning")
namein=str((date.today()-timedelta(days=2)).strftime('%d-%m-%Y'))
namefull= r"C:\Users\Aditya.gupta\Downloads\PKG C AT Tracker "+namein+r".xlsb"
loc_pkgC_AT=(namefull)
df_atC=pd.read_excel(loc_pkgC_AT,sheet_name="Master Sheet",skiprows=2)
##df_atC[r'Litup date'] = pd.TimedeltaIndex(df_atC[r'Litup date'], unit='d') + dt.datetime(1899, 12, 30)
df_atC[r'4-ATC released'] = pd.TimedeltaIndex(df_atC[r'4-ATC released'], unit='d') + dt.datetime(1899, 12, 30)
df_atC[r'T fiber document submitted'] = pd.TimedeltaIndex(df_atC[r'T fiber document submitted'], unit='d') + dt.datetime(1899, 12, 30)
df_atC.loc[df_atC["BOQ"] == "Submitted"] = np.NaN
df_atC[r'BOQ'] = pd.TimedeltaIndex(df_atC[r'BOQ'], unit='d') + dt.datetime(1899, 12, 30)
df_atC[r'Integrated PP cleared'] = pd.TimedeltaIndex(df_atC[r'Integrated PP cleared'], unit='d') + dt.datetime(1899, 12, 30)
##df_atC[r'BOQ'] = pd.TimedeltaIndex(df_atC[r'BOQ'], unit='d',errors='coerce') + dt.datetime(1899, 12, 30)
print("Got package C at info now getting doc's sub and BoQ status")
##done



##Pkg B DPR data manipulation
df2=df_dprB['T&D'].groupby(df_dprB[r' Date of Activity']).sum().reset_index()
df4=df_dprB['DRT'].groupby(df_dprB[r' Date of Activity']).sum().reset_index()
df3=df_dprB['Blowing'].groupby(df_dprB[r' Date of Activity']).sum().reset_index()
df_final=pd.concat([df2,df4['DRT'],df3['Blowing']], axis=1)
df_final=df_final.set_index(r' Date of Activity')
df_final= df_final.sort_index()
df_2 = df_final.loc[first_date :today_date].reset_index()
df_21 = df_final.loc[first_date1 :today_date].reset_index()

##Pkg C DPR data manipulation
df7=df_dprC['T&D'].groupby(df_dprC[r' Date of Activity']).sum().reset_index()
df8=df_dprC['DRT'].groupby(df_dprC[r' Date of Activity']).sum().reset_index()
df9=df_dprC['Blowing'].groupby(df_dprC[r' Date of Activity']).sum().reset_index()
df_final1=pd.concat([df7,df8['DRT'],df9['Blowing']], axis=1)
df_final1=df_final1.set_index(r' Date of Activity')
df_final1= df_final1.sort_index()
df_3 = df_final1.loc[first_date :today_date].reset_index()
df_31 = df_final1.loc[first_date1 :today_date].reset_index()

##Data input to excel (PkgB DPR)
print("Data entering in xl for PkgB DPR")
today_TnD_B= (df_2.loc[df_2[' Date of Activity']==today_date,'T&D']).sum()
month_TnD_B= (df_2.loc[(df_2[' Date of Activity']>=first_date) & (df_2[' Date of Activity']<=today_date),'T&D'].sum())
today_Blow_B= (df_2.loc[df_2[' Date of Activity']==today_date,'Blowing']).sum()
month_Blow_B= (df_2.loc[(df_2[' Date of Activity']>=first_date) & (df_2[' Date of Activity']<=today_date),'Blowing'].sum())
sheet['E2']=round(month_TnD_B)
sheet['G2']=round(today_TnD_B)
sheet['E3']=round(month_Blow_B)
sheet['G3']=round(today_Blow_B)


##Data input to excel (PkgC DPR)
print("Data entering in xl for PkgC DPR")
today_TnD_C= (df_3.loc[df_3[' Date of Activity']==today_date,'T&D']).sum()
month_TnD_C= (df_3.loc[(df_3[' Date of Activity']>=first_date) & (df_3[' Date of Activity']<=today_date),'T&D'].sum())
today_Blow_C= (df_3.loc[df_3[' Date of Activity']==today_date,'Blowing']).sum()
month_Blow_C= (df_3.loc[(df_3[' Date of Activity']>=first_date) & (df_3[' Date of Activity']<=today_date),'Blowing'].sum())
sheet['E10']=round(month_TnD_C)
sheet['G10']=round(today_TnD_C)
sheet['E11']=round(month_Blow_C)
sheet['G11']=round(today_Blow_C)


##Extracting info from package B POP and writing in excel
print("Data entering in xl for GPLit B")
today_gpPOPB= (df_POPB['GP Lit Up Date']==today_date).sum()
month_gpPOPB_yest= ((df_POPB['GP Lit Up Date']>=first_date) & (df_POPB['GP Lit Up Date']<=yesterday_date)).sum()
month_gpPOPB_today= ((df_POPB['GP Lit Up Date']>=first_date) & (df_POPB['GP Lit Up Date']<=today_date)).sum()
sheet['E4']=month_gpPOPB_today
sheet['G4']=month_gpPOPB_today-old_GPlitB

##Extracting info from package C POP and writing in excel
print("Data entering in xl for GPLit C")
today_gpPOPC= (df_POPC['GP Lit Up Date']==today_date).sum()
month_gpPOPC_yest= ((df_POPC['GP Lit Up Date']>=first_date) & (df_POPC['GP Lit Up Date']<=yesterday_date)).sum()
month_gpPOPC_today= ((df_POPC['GP Lit Up Date']>=first_date) & (df_POPC['GP Lit Up Date']<=today_date)).sum()
sheet['E12']=month_gpPOPC_today
sheet['G12']=month_gpPOPC_today-old_GPlitC
## None and Done



##Extracting info from package B at and writing in excel
print("Data entering in xl for GPAT B")
today_gpATB= (df_atB['Common ATC\n4 Dt']==today_date).sum()
month_gpATB_yest= ((df_atB['Common ATC\n4 Dt']>=first_date) & (df_atB['Common ATC\n4 Dt']<=yesterday_date)).sum()
month_gpATB_today= ((df_atB['Common ATC\n4 Dt']>=first_date) & (df_atB['Common ATC\n4 Dt']<=today_date)).sum()
month_gpPPB_today= ((df_atB['Common PPs cleared  Date']>=first_date) & (df_atB['Common PPs cleared  Date']<=today_date)).sum()
sheet['E6']=month_gpATB_today
sheet['E5']=month_gpPPB_today
sheet['G5']=month_gpPPB_today-old_PPClearB
sheet['G6']=month_gpATB_today-old_GPATB

print("Data entering in xl for GPAT B")
today_docB= (df_docB['Submit Date to T_Fiber']==today_date).sum()
month_docB_today= ((df_docB['Submit Date to T_Fiber']>=first_date) & (df_docB['Submit Date to T_Fiber']<=today_date)).sum()
month_boqB_today= ((df_docB['BOQ Approval Status']>=first_date) & (df_docB['BOQ Approval Status']<=today_date)).sum()
sheet['E7']=month_docB_today
sheet['G7']=month_docB_today-old_docB
sheet['E8']=month_boqB_today
sheet['G8']=month_boqB_today-old_boqB

print("Data entering in xl for GPAT B and Doc B and BoQ B")



##Extracting info from package C AT and writing in excel
print("Data entering in xl for GPAT C")
today_gpATC= (df_atC['4-ATC released']==today_date).sum()
month_gpATC_yest= ((df_atC['4-ATC released']>=first_date) & (df_atC['4-ATC released']<=yesterday_date)).sum()
month_gpATC_today= ((df_atC['4-ATC released']>=first_date) & (df_atC['4-ATC released']<=today_date)).sum()
print("Still inserting data in XL")
today_docC= (df_atC['T fiber document submitted']==today_date).sum()
month_docC_today= ((df_atC['T fiber document submitted']>=first_date) & (df_atC['T fiber document submitted']<=today_date)).sum()
today_boqC= (df_atC['BOQ']==today_date).sum()
month_boqC_today= ((df_atC['BOQ']>=first_date) & (df_atC['BOQ']<=today_date)).sum()
month_PPC_today= ((df_atC['Integrated PP cleared']>=first_date) & (df_atC['Integrated PP cleared']<=today_date)).sum()
sheet['E13']=month_PPC_today
sheet['G13']=month_PPC_today-old_PPClearC
sheet['E14']=month_gpATC_today
sheet['G14']=month_gpATC_today-old_GPATC
sheet['E15']=month_docC_today
sheet['G15']=month_docC_today-old_docC
sheet['E16']=month_boqC_today
sheet['G16']=month_boqC_today-old_boqC



namein=str((date.today()-timedelta(days=2)).strftime('%d-%b-%y'))
namefull= r"XL_dated_"+namein+r".xlsx"
namefull1= r"summ_dated_"+namein+".bmp"
xfile.save(namefull)
excel2img.export_img(namefull, namefull1, "", "Sheet1!A1:H16")


print("Plotting starts")
##plotting and saving the file both B and C DPR's plot
w=0.5
x=df_21[r' Date of Activity'].dt.strftime('%d-%b')
x1=df_21['T&D']
x2=df_21['DRT']
x3=df_31['T&D']
x4=df_31['DRT']
print(np.shape(x))
print(np.shape(df_01))
bar1=np.arange(len(x))
bar2= [i+w for i in bar1]
bar3= [i+w/2 for i in bar1]

##pkg B DPRs plot coming
plt.figure(1,figsize=(11, 5))
plt.subplots_adjust(bottom=0.23)
plt.grid(b=True, axis='y',zorder=0)
plt.bar(bar1,x1,w,label='T&D',color='blue',zorder=3)


plt.bar(bar2,x2,w,label='DRT',color='orange',zorder=3)


plt.plot(bar3,df_21['Blowing'], color='grey',label='Blowing',zorder=3)
plt.plot(bar3,df_21['Blowing'], color='grey',marker='o',zorder=3)
plt.legend(loc='upper left')

plt.text(bar3[0]-1.5,df_01['T&D'][0]+0.2 , "T&D Plan", horizontalalignment='left', size=5, color='black')
plt.text(bar3[0]-1.5,df_01['DRT'][0]+0.2 , "DRT Plan", horizontalalignment='left', size=5, color='black')
plt.text(bar3[0]-1.5,df_01['Blowing'][0]+0.2 , "Blowing Plan", horizontalalignment='left', size=5, color='black')

plt.plot(bar3,df_01['T&D'],color='blue',zorder=2,linewidth=0.750)
plt.plot(bar3,df_01['DRT'],color='orange',zorder=2,linewidth=0.750)
plt.plot(bar3,df_01['Blowing'],color='grey',zorder=2,linewidth=0.750)


plt.xticks(bar1+w/2,x, rotation=90)
plt.xlabel('Date ')
plt.ylabel('Progress in Kms.')
plt.legend()
plt.title('Execution Progress (Pkg. B)')
namein=str((date.today()-timedelta(days=2)).strftime('%d-%b-%y'))
namefull= r"pkgB_"+namein
plt.savefig(namefull)

## for pkg C DPR chart
plt.figure(2,figsize=(11, 5))
plt.subplots_adjust(bottom=0.23)
plt.grid(b=True, axis='y',zorder=0)
plt.bar(bar1,x3,w,label='T&D',color='blue',zorder=3)


plt.bar(bar2,x4,w,label='DRT',color='orange',zorder=3)


plt.plot(bar3,df_31['Blowing'], color='grey',label='Blowing',zorder=3)
plt.plot(bar3,df_31['Blowing'], color='grey',marker='o',zorder=3)

plt.text(bar3[0]-1.5,df_001['T&D'][0]+0.2 , "T&D Plan", horizontalalignment='left', size=5, color='black')
plt.text(bar3[0]-1.5,df_001['DRT'][0]+0.2 , "DRT Plan", horizontalalignment='left', size=5, color='black')
plt.text(bar3[0]-1.5,df_001['Blowing'][0]+0.2 , "Blowing Plan", horizontalalignment='left', size=5, color='black')

plt.plot(bar3,df_001['T&D'],color='blue',zorder=2,linewidth=0.750)
plt.plot(bar3,df_001['DRT'],color='orange',zorder=2,linewidth=0.750)
plt.plot(bar3,df_001['Blowing'],color='grey',zorder=2,linewidth=0.750)
plt.legend(loc='upper left')

plt.xticks(bar1+w/2,x, rotation=90)
plt.xlabel('Date ')
plt.ylabel('Progress in Kms.')
plt.legend()
plt.title('Execution Progress (Pkg. C)')
namein=str((date.today()-timedelta(days=2)).strftime('%d-%b-%y'))
namefull= r"pkgC_"+namein
plt.savefig(namefull)


